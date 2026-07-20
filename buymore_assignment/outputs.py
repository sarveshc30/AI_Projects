"""Output writers: append-only Excel row (pandas) and the per-brand Markdown report."""

import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from analysis import render_category_tree
from logger import log, warn

EXCEL_PATH = "output.xlsx"


def _leaf_paths(nodes, prefix=()):
    """Every root-to-leaf breadcrumb path in the category tree."""
    paths = []
    for node in nodes:
        path = (*prefix, node.name)
        if node.children:
            paths += _leaf_paths(node.children, path)
        else:
            paths.append(path)
    return paths


def _root_categories(nodes):
    """Every top-level category the brand sells under, busiest first."""
    return [node.name for node in nodes]


def _leaf_categories(nodes):
    """Every deepest-level sub-category, deduplicated.

    The same leaf can be reached from more than one root (e.g. "Lotions" under both
    Beauty and Baby Products), so identical names collapse to one entry while keeping the
    tree's count-sorted order.
    """
    names = []
    for path in _leaf_paths(nodes):
        if path[-1] not in names:
            names.append(path[-1])
    return names

COLUMNS = [
    "Brand Name",
    "Category",
    "Sub-Category",
    "Portals Live",
    "Check Ratings",
    "Sellers Name",
    "Running Ads",
    "Listing Quality",
    "Weakness Report",
]

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True, size=11)


def _row_from_report(report):
    # Every top-level category and every deepest sub-category the brand was seen in, one
    # per line. The tree can be empty if no breadcrumbs were captured, so fall back to the
    # single resolved pair rather than writing a blank cell.
    roots = _root_categories(report.category_tree) or (
        [report.category] if report.category else []
    )
    leaves = _leaf_categories(report.category_tree) or (
        [report.sub_category] if report.sub_category else []
    )

    return {
        "Brand Name": report.brand_name,
        "Category": "\n".join(roots),
        "Sub-Category": "\n".join(leaves),
        "Portals Live": ", ".join(report.portals_live),
        "Check Ratings": "\n".join(report.check_ratings_urls),
        "Sellers Name": ", ".join(report.sellers),
        "Running Ads": report.running_ads_label,
        "Listing Quality": report.audit.score,
        "Weakness Report": "\n".join(f"- {b}" for b in report.weaknesses),
    }


def _style_sheet(path):
    """Style headers and size columns so the file is readable when opened."""
    workbook = load_workbook(path)
    sheet = workbook.active
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # B/C stack one category per line, so they need more room than a single value would.
    widths = {"A": 18, "B": 24, "C": 30, "D": 22, "E": 55, "F": 32, "G": 12, "H": 14, "I": 80}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)


def write_excel(report, path=EXCEL_PATH):
    """Append one row per brand. Existing rows are read back and preserved."""
    row = _row_from_report(report)
    try:
        if os.path.exists(path):
            existing = pd.read_excel(path)
            # Reindex guards against an older file written with a different column set.
            frame = pd.concat([existing, pd.DataFrame([row])], ignore_index=True)
            frame = frame.reindex(columns=COLUMNS)
        else:
            frame = pd.DataFrame([row], columns=COLUMNS)

        frame.to_excel(path, index=False, engine="openpyxl")
        _style_sheet(path)
        log("excel", f"saved row {len(frame) + 1} to {path}")
        return True
    except PermissionError:
        warn("excel", f"{path} is open in another program - close it and re-run to append this row")
    except Exception as e:
        warn("excel", f"failed to write {path}: {e}")
    return False


def _markdown_body(report):
    lines = [f"# Brand Intelligence Report: {report.brand_name}", ""]

    lines += [
        "## Brand Overview",
        "",
        f"- **Found on Amazon.in:** {'Yes' if report.found_on_amazon else 'No'}",
        f"- **Portals live:** {', '.join(report.portals_live) if report.portals_live else 'None confirmed'}",
        f"- **Category:** {report.category or 'Not determined'}",
        f"- **Sub-Category:** {report.sub_category or 'Not determined'}",
        "",
    ]

    lines += ["## Category Footprint", ""]
    if report.category_tree:
        lines += [
            "Full breadcrumb hierarchy merged across every scraped product page. "
            "The number in brackets is how many of the scraped products sit under that node.",
            "",
            "```",
        ]
        lines += render_category_tree(report.category_tree)
        lines += ["```", ""]

        lines += ["**Distinct category paths:**", ""]
        for path in _leaf_paths(report.category_tree):
            lines.append(f"- {' > '.join(path)}")
    else:
        lines.append("No category breadcrumbs were captured from the scraped product pages.")
    lines.append("")

    lines += ["## Product Catalog", ""]
    if report.products:
        lines += ["| # | Product | Ratings | Sponsored | URL |", "|---|---|---|---|---|"]
        for i, product in enumerate(report.products, 1):
            title = (product.title or "Unknown")[:70]
            # A per-product "No" is only meaningful if ads were being served at all. With
            # no fill on the page, every row would otherwise read No while the section
            # below correctly reports Unknown.
            if product.is_sponsored:
                ad_flag = "Yes"
            elif report.ad_slots_served:
                ad_flag = "No"
            else:
                ad_flag = "Unknown"
            lines.append(
                f"| {i} | {title} | {product.rating_count:,} | {ad_flag} | {product.url} |"
            )
    else:
        lines.append("No product URLs were discovered.")
    lines.append("")

    lines += ["## Marketplace Intelligence", ""]
    lines += [
        f"- **Unique sellers ({len(report.sellers)}):** "
        f"{', '.join(report.sellers) if report.sellers else 'None extracted'}",
        f"- **Running sponsored ads:** {report.running_ads_label}",
        f"- **Products with 100+ ratings:** {len(report.check_ratings_urls)}",
        f"- **Sponsored listings on the search page:** {len(report.sponsored_products)}",
    ]
    lines.append("")

    lines += ["### Sponsored Ad Presence", ""]
    if report.sponsored_products:
        lines += [
            f"These {len(report.sponsored_products)} "
            f"{report.brand_name} product(s) carried a **Sponsored** tag on the "
            f'Amazon.in search results page for "{report.brand_name}", so the brand is '
            "paying for placement against its own brand query:",
            "",
        ]
        lines += [f"{i}. {name}" for i, name in enumerate(report.sponsored_products, 1)]
        lines += [
            "",
            "_Detected from the search results page - sponsored status is a property of "
            "how a listing appears in search, not of the product detail page. Tiles are "
            "matched on the ad-feedback label, a literal Sponsored badge, or an "
            "`/sspa/click` tracking redirect._",
        ]
    elif report.ad_slots_served:
        lines += [
            f"**No** - no sponsored tiles for {report.brand_name}, although other brands' "
            "ads did serve on the same search page. The ad slots were live and this brand "
            "was absent from them, so it is not bidding on its own brand query.",
        ]
    else:
        lines += [
            "**Unknown** - Amazon served no sponsored slots at all on this search page, "
            "for any brand.",
            "",
            "_Sponsored placements are filled client-side by a separate ad call that "
            "routinely returns nothing for an automated or cookie-less session, while the "
            "organic results still render normally. That makes an empty result "
            "indistinguishable from a brand that genuinely does not advertise, so this is "
            "recorded as Unknown rather than No. Re-run headed to get a usable answer._",
        ]
    lines.append("")

    audit = report.audit
    top = report.top_product
    lines += ["## Listing Quality Audit", ""]
    lines.append(f"**Score: {audit.score}/10**")
    lines.append("")
    if top:
        lines.append(f"Audited listing: [{(top.title or top.url)[:80]}]({top.url})")
        lines.append("")
    lines.append(f"_{audit.remark}_")
    lines += [
        "",
        "| Dimension | Assessment | Parameters |",
        "|---|---|---|",
        f"| Title Quality | {audit.title_quality} | {audit.title_word_count} words; "
        f"brand in title: {audit.title_includes_brand}; sub-category in title: {audit.title_includes_subcategory} |",
        f"| Visual Quality | {audit.visual_quality} | {audit.image_count} images |",
        f"| Content Richness | {audit.content_richness} | bullets: {audit.has_bullets}; "
        f"description: {audit.has_description}; specifications: {audit.has_specifications}; "
        f"A+ content: {audit.has_aplus_content} |",
        f"| Data Accuracy | {audit.data_accuracy} | pack size stated: {audit.pack_size_stated}; "
        f"ingredients listed: {audit.ingredients_listed} |",
        f"| Social Proof | {audit.social_proof} | {audit.star_rating or 'n/a'} stars from "
        f"{audit.rating_count:,} ratings |",
        "",
    ]

    lines += ["## Weakness Report", ""]
    if report.weaknesses:
        lines += [f"{i}. {b}" for i, b in enumerate(report.weaknesses, 1)]
    else:
        lines.append("No weaknesses generated.")
    lines.append("")

    return "\n".join(lines)


def write_markdown(report, directory="."):
    """Write <brand_name>.md next to the Excel file."""
    safe_name = "".join(c for c in report.brand_name if c.isalnum() or c in " -_").strip()
    path = os.path.join(directory, f"{safe_name or 'brand'}.md")
    try:
        with open(path, "w", encoding="utf-8") as handle:
            handle.write(_markdown_body(report))
        log("markdown", f"saved {path}")
        return True
    except Exception as e:
        warn("markdown", f"failed to write {path}: {e}")
    return False
