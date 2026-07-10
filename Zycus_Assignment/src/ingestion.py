"""Config-driven ingestion of 3-sheet project plan workbooks into canonical records.

No source-file-specific branching lives here beyond generic fallback heuristics for an
unrecognized file. Per-file column quirks are resolved entirely via config/column_mapping.yaml.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Optional

import openpyxl
import yaml

CONFIG_PATH = Path(__file__).resolve().parent.parent / "config" / "column_mapping.yaml"


def _normalize_key(name: str) -> str:
    return re.sub(r"[\s_]+", "_", name.strip().lower())


def load_column_mapping(config_path: Path = CONFIG_PATH) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


class IngestionResult:
    """Raw (uncleaned) canonical records straight out of the workbook."""

    def __init__(
        self,
        project_file: str,
        task_rows: list[dict[str, Any]],
        comments: list[dict[str, Any]],
        summary: dict[str, Any],
        data_gaps: list[str],
    ):
        self.project_file = project_file
        self.task_rows = task_rows
        self.comments = comments
        self.summary = summary
        self.data_gaps = data_gaps


def _resolve_file_profile(mapping: dict, filename: str) -> Optional[dict]:
    key = _normalize_key(Path(filename).name)
    return mapping["file_profiles"].get(key)


def _pick_task_sheet(wb, profile: Optional[dict]) -> str:
    if profile and profile.get("task_sheet") in wb.sheetnames:
        return profile["task_sheet"]
    for name in wb.sheetnames:
        if name.strip().lower() not in ("comments", "summary"):
            return name
    raise ValueError(f"No task sheet found among sheets: {wb.sheetnames}")


def _pick_sheet(wb, profile: Optional[dict], profile_key: str, default_name: str) -> Optional[str]:
    if profile and profile.get(profile_key) in wb.sheetnames:
        return profile[profile_key]
    for name in wb.sheetnames:
        if name.strip().lower() == default_name:
            return name
    return None


def ingest_workbook(file_path: str | Path, mapping: Optional[dict] = None) -> IngestionResult:
    file_path = Path(file_path)
    mapping = mapping or load_column_mapping()
    wb = openpyxl.load_workbook(file_path, data_only=True)
    profile = _resolve_file_profile(mapping, file_path.name)
    file_key = _normalize_key(file_path.name)

    data_gaps: list[str] = []
    if profile is None:
        data_gaps.append(
            f"'{file_path.name}' does not match a known profile in column_mapping.yaml; "
            "falling back to heuristic sheet/column detection. Add a profile for this file "
            "to config/column_mapping.yaml for reliable results."
        )

    task_sheet_name = _pick_task_sheet(wb, profile)
    comments_sheet_name = _pick_sheet(wb, profile, "comments_sheet", "comments")
    summary_sheet_name = _pick_sheet(wb, profile, "summary_sheet", "summary")

    task_rows = _ingest_task_sheet(wb[task_sheet_name], mapping, file_key, data_gaps)

    comments = (
        _ingest_comments_sheet(wb[comments_sheet_name], mapping)
        if comments_sheet_name
        else []
    )
    if not comments_sheet_name:
        data_gaps.append("No Comments sheet found; stakeholder sentiment cannot be computed.")
    elif not comments:
        data_gaps.append(
            "Comments sheet is present but contains zero usable comments; stakeholder "
            "sentiment defaults to neutral rather than being fabricated."
        )

    summary = (
        _ingest_summary_sheet(wb[summary_sheet_name], mapping)
        if summary_sheet_name
        else {}
    )
    if not summary_sheet_name:
        data_gaps.append(
            "No Summary sheet found; project-level rollups (Today's Date, Schedule Health, "
            "milestone/phase counts) are unavailable."
        )

    return IngestionResult(file_path.name, task_rows, comments, summary, data_gaps)


def _ingest_task_sheet(
    ws, mapping: dict, file_key: str, data_gaps: list[str]
) -> list[dict[str, Any]]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [c.value for c in header_row]
    header_index = {h: i for i, h in enumerate(headers) if h is not None}

    col_map = mapping["task_sheet_column_map"]
    known_file_keys: set[str] = set()
    for per_file in col_map.values():
        known_file_keys.update(per_file.keys())
    use_known_profile = file_key in known_file_keys

    field_source_header: dict[str, Optional[str]] = {}
    for canonical_field, per_file in col_map.items():
        if use_known_profile:
            source_header = per_file.get(file_key)
        else:
            # Unrecognized file: accept any configured header name that actually exists here.
            source_header = next(
                (cand for cand in per_file.values() if cand and cand in header_index), None
            )
        field_source_header[canonical_field] = source_header

    fallback_field = mapping.get("level_fallback", {}).get("fallback_field")
    level_uses_fallback = field_source_header.get("level") is None and fallback_field
    if level_uses_fallback:
        data_gaps.append(
            f"'Level' column absent from task sheet; using '{fallback_field}' values as a "
            "hierarchy-depth proxy (empirically comparable distribution, but treat with caution)."
        )

    missing_fields = sorted(
        f
        for f, h in field_source_header.items()
        if h is None and not (f == "level" and level_uses_fallback)
    )
    if missing_fields:
        data_gaps.append(
            "Task sheet has no matching source column for: "
            + ", ".join(missing_fields)
            + ". These fields are set to None, not guessed or defaulted."
        )

    rows: list[dict[str, Any]] = []
    excel_row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        excel_row_num += 1
        if all(v is None for v in row):
            continue
        record: dict[str, Any] = {"_row_id": excel_row_num, "_source_file": file_key}
        for canonical_field, source_header in field_source_header.items():
            if source_header is None:
                record[canonical_field] = None
                continue
            idx = header_index.get(source_header)
            record[canonical_field] = row[idx] if idx is not None and idx < len(row) else None
        if level_uses_fallback:
            record["level"] = record.get(fallback_field)
        rows.append(record)

    return rows


def _ingest_comments_sheet(ws, mapping: dict) -> list[dict[str, Any]]:
    columns = mapping["comments_sheet_format"]["columns"]
    comments: list[dict[str, Any]] = []
    for row in ws.iter_rows(values_only=True):
        if row is None or all(v is None for v in row):
            continue
        record = {col: (row[i] if i < len(row) else None) for i, col in enumerate(columns)}
        if record.get("comment_text"):
            comments.append(record)
    return comments


def _ingest_summary_sheet(ws, mapping: dict) -> dict[str, Any]:
    aliases = mapping["summary_key_aliases"]
    summary: dict[str, Any] = {}
    unmapped: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        raw_key = row[0]
        value = row[1] if len(row) > 1 else None
        canonical_key = aliases.get(raw_key)
        if canonical_key:
            summary[canonical_key] = value
        else:
            unmapped[raw_key] = value
    if unmapped:
        summary["_unmapped"] = unmapped
    return summary
